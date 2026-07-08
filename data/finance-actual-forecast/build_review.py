# -*- coding: utf-8 -*-
"""判定シート(hantei_sheet.xlsx)を生成。
A=個別判定(銀行要確認＋カード大口5万↑/記録口座・資金フロー付)、B=カード小口パターン、
C=前提と論点、D=口座間資金移動(出金⇄入金の自動マッチング)。
対象＝利用日2025/2026（全期間累計の論点はCに記載）。"""
import csv, os, re, unicodedata
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

BASE = os.path.dirname(os.path.abspath(__file__))
rows = list(csv.DictReader(open(os.path.join(BASE, "ledger_master.csv"), encoding="utf-8-sig")))

def yr(r): return r["use_date"][:4]
def amt(r): return int(r["amount"])

CHOICES = ["法人経費", "個人費用(生活費)", "役員報酬", "役員貸付(法人→個人)",
           "役員借入(個人→法人)", "売上", "不課税・除外(振替/返済/二重)", "保留"]

# 口座の短縮名
SRC_SHORT = {
    "住信SBI(法人)": "SBI法人", "住信SBI(個人)": "SBI個人",
    "SMBC(法人)": "SMBC法人", "SMBC(個人_ときわ台)": "SMBC個人(ときわ台)",
    "SMBC(個人_新宿西口)": "SMBC個人(新宿西口)", "楽天銀行(個人)": "楽天個人",
}
def short(src): return SRC_SHORT.get(src, src)

def clean_name(n):
    n = unicodedata.normalize("NFKC", n)
    n = re.sub(r"[（(][^）)]*[）)]", "", n)
    n = re.sub(r"[0-9]{3,}", "", n)
    n = n.replace("＊", "").replace("*", "")
    n = re.sub(r"\s+", "", n)
    return n[:26] or "(名称なし)"

# 相手方エンティティ推定（資金フロー表示用）
def counterparty(name):
    u = clean_name(name)
    u = re.sub(r"^(振込サービス|振込専用|振込|口座振替|パソコン振込|パソコン振替)", "", u).strip() or u
    if "イシダユウキ" in u: return "石田優輝(個人)"
    if "イテラ" in u: return "ITERRA(法人)"
    if "イシダカズナリ" in u: return "石田カズナリ(家族?)"
    if "イシダリナ" in u: return "石田リナ(家族?)"
    if "イシダ" in u: return "石田(個人/家族)"
    if any(k in u for k in ["セゾン", "ミツイスミトモ", "ニコス", "ＤＣ", "DC", "オリコ", "オリエント", "アプラス", "クレデイ"]): return "カード会社"
    if any(k in u for k in ["ヤチン", "ジエイリース", "リンクス", "ノムラフドウサン", "フドウサン"]): return "家賃/不動産先"
    if "コウセイロウドウシヨウ" in u or "ネンキン" in u: return "年金事務所"
    if "ゼイリシ" in u: return "税理士"
    if "ポイントキコウ" in u: return "ポイント機構(取引先)"
    if "グロ" in u: return "グローシング(取引先?)"
    if "コウムブ" in u: return "東京公務部"
    if "ATM" in u or "支払機" in u or "預金機" in u or "カ-ド出金" in u or "カード出金" in u: return "現金(ATM)"
    return u

def flow_arrow(src, flow, name):
    cp = counterparty(name)
    s = short(src)
    return f"{s} → {cp}" if flow == "出金" else f"{cp} → {s}"

# ===== 集計（対象＝2025/2026） =====
TARGET = ("2025", "2026")
BANK_THRESH = 10000

# A銀行: (source, owner, flow, cleanname)
bankA = defaultdict(lambda: defaultdict(int))
for r in rows:
    if r["type"] == "bank" and r["segment"] == "要確認" and yr(r) in TARGET:
        k = (r["source"], r["owner"], r["flow"], clean_name(r["name"]))
        bankA[k][yr(r)] += amt(r)
        bankA[k]["_acct"] = r["account"]

# Aカード大口・Bグループ
def card_group(a):
    if "Amazon" in a: return "Amazon(品目混在・要注文履歴)"
    if any(x in a for x in ["宿泊", "旅行", "航空", "海外旅行保険", "レンタカー(海外)"]): return "旅行・宿泊(国内/海外)"
    if "飲食" in a: return "飲食店(会議/接待/私的の別)"
    if "交通系IC" in a or "PASMO" in a or "Suica" in a: return "交通系IC(PASMO/Suica)"
    if "Google課金" in a or "スマホ端末" in a: return "Google課金・スマホ端末"
    if "YouTube" in a: return "サブスク(YouTube等・私的性)"
    if a.startswith("（個人候補）"): return "生活費・物販(衣服/家電/食料/百貨店等)"
    if "未分類" in a: return "未分類(その他カード)"
    return "その他カード要確認"

tmp = defaultdict(lambda: [0, 0, "", "", "", ""])  # name->[y25,y26,acct,owner,seg,src]
for r in rows:
    if r["type"] == "card" and r["segment"] in ("要確認", "個人候補") and yr(r) in TARGET:
        t = tmp[r["name"][:34]]
        if yr(r) == "2025": t[0] += amt(r)
        else: t[1] += amt(r)
        t[2] = r["account"]; t[3] = r["owner"]; t[4] = r["segment"]; t[5] = r["source"]

THRESH = 50000
cardA = []
groupB = defaultdict(lambda: [0, 0, 0, set()])
bank_small = defaultdict(lambda: [0, 0, 0, set()])
for name, (y25, y26, acct, owner, seg, src) in tmp.items():
    if (y25 + y26) >= THRESH and seg == "要確認":
        cardA.append((name, y25, y26, acct, owner, src))
    else:
        g = card_group(acct); b = groupB[g]
        b[0] += y25; b[1] += y26; b[2] += 1
        if len(b[3]) < 4: b[3].add(name[:16])

# ===== スタイル =====
wb = openpyxl.Workbook()
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdrfill = PatternFill("solid", fgColor="305496"); hdrfont = Font(color="FFFFFF", bold=True)
inputfill = PatternFill("solid", fgColor="FFF2CC")
flowfill = PatternFill("solid", fgColor="E2EFDA")
secfont = Font(bold=True, size=12, color="305496")
wrap = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c); cell.fill = hdrfill; cell.font = hdrfont
        cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = border

# ===== Sheet A（取引明細を1件ずつ）=====
wsA = wb.active; wsA.title = "A_個別判定"
dvA = DataValidation(type="list", formula1='"' + ",".join(CHOICES) + '"', allow_blank=True)
wsA.add_data_validation(dvA)
wsA["A1"] = "【A】個別判定（銀行の要確認を1取引ずつ表示）／緑＝資金の流れ・黄＝判定入力"; wsA["A1"].font = secfont
wsA["A2"] = "年集計ではなく取引明細単位。日付・金額を見て判定できる。『資金の流れ』はFrom→To(推定)。利用日ベース・円。"; wsA["A2"].font = Font(size=9, color="808080")
hdr = ["利用日", "名義", "記録口座", "入出金", "相手先/摘要", "資金の流れ(推定)", "推定科目", "金額", "【判定】", "【メモ】"]
r0 = 4
for j, h in enumerate(hdr, 1): wsA.cell(r0, j, h)
style_header(wsA, r0, len(hdr))
widths = [11, 5, 16, 6, 26, 28, 28, 12, 16, 22]
for j, w in enumerate(widths, 1): wsA.column_dimensions[chr(64 + j)].width = w

# 銀行 要確認を取引明細単位で（2025/2026）、金額降順
bank_tx = [r for r in rows if r["type"] == "bank" and r["segment"] == "要確認" and yr(r) in TARGET]
bank_tx.sort(key=lambda r: -amt(r))
rr = r0 + 1
for r in bank_tx:
    arrow = flow_arrow(r["source"], r["flow"], r["name"])
    vals = [r["use_date"], r["owner"], short(r["source"]), r["flow"],
            clean_name(r["name"]), arrow, r["account"], amt(r)]
    for j, v in enumerate(vals, 1):
        cell = wsA.cell(rr, j, v); cell.border = border
        if j == 8: cell.number_format = "#,##0"
        if j == 6: cell.fill = flowfill
    jc = wsA.cell(rr, 9); jc.fill = inputfill; jc.border = border; dvA.add(jc)
    wsA.cell(rr, 10).fill = inputfill; wsA.cell(rr, 10).border = border
    rr += 1
wsA.cell(rr + 1, 1, f"銀行要確認 {len(bank_tx)}件・計 {sum(amt(r) for r in bank_tx):,}円（カードの要確認はBシート）").font = Font(bold=True)
wsA.freeze_panes = "A5"

# カード大口はBシートへ回す（個別行として）
for name, y25, y26, acct, owner, src in cardA:
    g = f"[カード大口] {name[:20]}"; b = groupB[g]
    b[0] += y25; b[1] += y26; b[2] += 1; b[3].add(short(src))

# ===== Sheet D：口座間資金移動（出金⇄入金マッチング） =====
wsD = wb.create_sheet("D_口座間資金移動")
wsD["A1"] = "【D】口座間の資金移動（自動マッチング：別口座の 出金 と 入金 で 同額×日付±5日 を対応付け）"; wsD["A1"].font = secfont
wsD["A2"] = "法人→個人 等の実際の振替を可視化。推定一致のため要確認。同額の偶然一致が混じる場合あり。"; wsD["A2"].font = Font(size=9, color="808080")
hdrD = ["出金日", "From口座", "From摘要", "入金日", "To口座", "To摘要", "金額", "方向(名義)"]
r0 = 4
for j, h in enumerate(hdrD, 1): wsD.cell(r0, j, h)
style_header(wsD, r0, len(hdrD))
for j, w in enumerate([11, 16, 24, 11, 16, 24, 12, 14], 1): wsD.column_dimensions[chr(64 + j)].width = w

# 全銀行取引（除外も含めて＝振替は除外扱いのものが多いため）を出金/入金に分ける
debits = [r for r in rows if r["type"] == "bank" and r["flow"] == "出金" and yr(r) in TARGET]
credits = [r for r in rows if r["type"] == "bank" and r["flow"] == "入金" and yr(r) in TARGET]
def d2i(s):
    y, m, dd = s.split("/"); return int(y) * 372 + int(m) * 31 + int(dd)
used = set()
pairs = []
for de in sorted(debits, key=lambda r: -amt(r)):
    if amt(de) < 10000:  # 1万未満の偶然一致は除く
        continue
    best = None
    for i, cr in enumerate(credits):
        if i in used: continue
        if cr["source"] == de["source"]: continue
        if amt(cr) != amt(de): continue
        if abs(d2i(cr["use_date"]) - d2i(de["use_date"])) > 5: continue
        best = i; break
    if best is not None:
        used.add(best); cr = credits[best]
        owdir = f"{de['owner']}→{cr['owner']}"
        pairs.append((de["use_date"], short(de["source"]), clean_name(de["name"]),
                      cr["use_date"], short(cr["source"]), clean_name(cr["name"]),
                      amt(de), owdir))
pairs.sort(key=lambda x: -x[6])
rr = r0 + 1
for p in pairs:
    for j, v in enumerate(p, 1):
        cell = wsD.cell(rr, j, v); cell.border = border
        if j == 7: cell.number_format = "#,##0"
        if j == 8 and "法人→個人" in v: cell.fill = flowfill
    rr += 1
wsD.cell(rr + 1, 1, f"マッチ件数 {len(pairs)} 件・合計 {sum(p[6] for p in pairs):,} 円（うち法人→個人 {sum(p[6] for p in pairs if p[7]=='法人→個人'):,} 円）").font = Font(bold=True)
wsD.freeze_panes = "A5"

# ===== Sheet B =====
wsB = wb.create_sheet("B_カード小口パターン")
dvB = DataValidation(type="list", formula1='"' + ",".join(CHOICES) + '"', allow_blank=True)
wsB.add_data_validation(dvB)
wsB["A1"] = "【B】カード小口の要確認・個人候補：グループ単位で方針を決める"; wsB["A1"].font = secfont
wsB["A2"] = "グループごとに『基本どう扱うか』を黄色セルに。例外は備考か ledger_master.csv で個別指定。"; wsB["A2"].font = Font(size=9, color="808080")
hdrB = ["科目グループ", "2025計", "2026計", "件数", "代表例", "【基本の判定】", "【メモ/例外】"]
r0 = 4
for j, h in enumerate(hdrB, 1): wsB.cell(r0, j, h)
style_header(wsB, r0, len(hdrB))
for j, w in enumerate([34, 12, 12, 8, 40, 20, 30], 1): wsB.column_dimensions[chr(64 + j)].width = w
rr = r0 + 1
allB = dict(groupB); allB.update(bank_small)
for g, (y25, y26, cnt, ex) in sorted(allB.items(), key=lambda x: -(x[1][0] + x[1][1])):
    wsB.cell(rr, 1, g); wsB.cell(rr, 2, y25); wsB.cell(rr, 3, y26); wsB.cell(rr, 4, cnt)
    wsB.cell(rr, 5, " / ".join(sorted(ex)))
    for c in (2, 3): wsB.cell(rr, c).number_format = "#,##0"
    jc = wsB.cell(rr, 6); jc.fill = inputfill; dvB.add(jc)
    wsB.cell(rr, 7).fill = inputfill
    for c in range(1, 8): wsB.cell(rr, c).border = border
    rr += 1
wsB.freeze_panes = "A5"

# ===== Sheet C =====
wsC = wb.create_sheet("C_前提と論点")
wsC.column_dimensions["A"].width = 100
lines = [
    ("【判定の選択肢】", secfont),
    ("法人経費 / 個人費用(生活費) / 役員報酬 / 役員貸付(法人→個人) / 役員借入(個人→法人) / 売上 / 不課税・除外(振替・返済・二重計上) / 保留", None),
    ("", None),
    ("【口座間の資金移動（Dシート参照）】", secfont),
    ("・法人→個人 の振替は『役員報酬』か『役員貸付』かで税務が大きく変わる。Dシートで実際の口座ペアを確認し、Aで同じ判定を付ける。", None),
    ("・法人側『出金イシダユウキ』と個人側『入金イテラ』は同じ資金移動の表裏。同じ判定を（最終P&Lで二重計上しません）。", None),
    ("", None),
    ("【本人にしか分からない主要論点（Aシートで判定）】", secfont),
    ("① 役員振込（法人→個人）：全期間 出金約1,057万・入金約37万。役員報酬か役員貸付か。月いくらを報酬とみなすか。", None),
    ("② 社宅家賃 179,550/月（個人口座から MHFヤチンシュウノウ等で支払）：法人負担(地代家賃)か個人負担か。事業計画では役員社宅179,550で法人計上の前提。", None),
    ("③ SMBC法人 入金『トウキヨウコウムブ』199万：補助金/給付/還付/売上のどれか。", None),
    ("④ 人名振込（オカダ/ワタナベ ユウジ/サガワ/フクシマ/タナカ レイ 等）：外注費か給与か立替精算か。", None),
    ("   ※ワタナベ ユウジは法人で出金・楽天で入金＝双方向。関係の確認要。", None),
    ("⑤ ATM/支払機/カード出金（法人・個人）：使途。役員貸付か経費か生活費か。", None),
    ("⑥ グローシング社 入金：売上か。 / イテラ(自社間振替) は相殺でよいか。", None),
    ("", None),
    ("【Bシートで方針を決める小口（カード）】", secfont),
    ("・旅行・宿泊（国内/海外）/飲食店/Amazon/生活費・物販/交通系IC/Google課金 …各グループの原則的な扱い。", None),
    ("", None),
    ("【扱いが確定済み（参考）】", secfont),
    ("・カードの銀行引落（セゾン/三井住友=SMCC/NICOS=DC=JAL）は明細取込済→除外（二重計上回避）。", None),
    ("・オリコ(分割2契約)・Googleスマホ分割：品目明細なし→銀行/カード引落ベースで個人計上。総額は足さない。", None),
    ("・Paidy：2026支払いなし→対象外。", None),
]
rr = 1
for text, fnt in lines:
    c = wsC.cell(rr, 1, text)
    if fnt: c.font = fnt
    c.alignment = wrap; rr += 1

# シート順：A, D, B, C
wb.move_sheet("D_口座間資金移動", -(wb.sheetnames.index("D_口座間資金移動") - 1))
out = os.path.join(BASE, "hantei_sheet.xlsx")
wb.save(out)
print("生成:", out)
print("A明細:", len(bank_tx), "件 / Dマッチ:", len(pairs), "件 / Bグループ:", len(allB))
