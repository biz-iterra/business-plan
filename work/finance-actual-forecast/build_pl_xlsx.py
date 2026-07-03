# -*- coding: utf-8 -*-
"""法人・個人 収支 月次xlsx（事業計画フォーマット準拠の勘定科目）。
出力: 収支実績_法人個人_2025-2026.xlsx
各月次シート（法人2025/2026・個人2025/2026）に、親科目＋子内訳（サービス/契約単位）の行を表示。
- 売上：取引先別。販管費・個人支出は【固定費】【変動費】に分割、各科目に子内訳。
- ソフト=サービス別、通信/地代家賃/旅費=内訳、個人カード返済=オリコ契約A/B・アプラス・Paidy、Googleスマホ分割。
- ITA=個人の外貨積立(2026/7〜200USD=月25,000想定)。利用日ベース。Amazon・少額要確認は未反映。
末尾に統合サマリ。"""
import csv, os, re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
rows = list(csv.DictReader(open(os.path.join(BASE, "ledger_master.csv"), encoding="utf-8-sig")))
def yr(r): return r["use_date"][:4]
def mo(r):
    try: return int(r["use_date"][5:7])
    except: return 0
def A(r): return int(r["amount"])
MONTHS = list(range(1, 13))

# ---- 親科目マッピング ----
def kamoku_corp_exp(a):
    if "役員報酬" in a: return "役員報酬"
    if "外注費" in a: return "外注費"
    if "法定福利" in a or "社会保険" in a: return "法定福利費"
    if "給料" in a: return "給料手当"
    if "通信費" in a or "サーバー" in a or "IP電話" in a or "03plus" in a: return "通信費"
    if "ソフトウェア" in a: return "ソフトウェア利用料"
    if "支払報酬" in a or "税理士" in a: return "支払報酬料"
    if "地代家賃" in a: return "地代家賃"
    if "水道光熱" in a: return "水道光熱費"
    if any(k in a for k in ["旅費交通","宿泊","交通系","タクシー","レンタカー","空港","航空","シェアサイクル","LUUP","電車","Suica","PASMO","配車","旅行"]): return "旅費交通費"
    if "接待" in a: return "接待交際費"
    if "会議費" in a: return "会議費"
    if "広告" in a or "印刷" in a: return "広告宣伝費"
    if "消耗品" in a: return "消耗品費"
    if "新聞図書" in a: return "新聞図書費"
    if "租税" in a or "公課" in a: return "租税公課"
    if "手数料" in a: return "支払手数料"
    if "現金支出" in a: return "雑費（現金支出）"
    return "その他経費"

def kamoku_corp_inc(a):
    if "ポイント機構" in a: return "売上：ポイント機構（既存受託）"
    if "グローシング" in a: return "売上：グローシング"
    if "ピカソ" in a: return "売上：ピカソ"
    if "アイスリー" in a or "I3" in a: return "売上：アイスリー"
    if "ザ(" in a or "ザ（" in a: return "売上：ザ(株)"
    if "売上" in a: return "売上：その他"
    if "雑収入" in a or "利息" in a: return "雑収入"
    return "その他収入"

def kamoku_kojin(a):
    a = a.replace("（個人候補）", "")
    if "スマホ分割" in a or "SPLITIT" in a.upper(): return "スマホ分割（Google）"
    if "外貨積立" in a: return "外貨積立（USD貯蓄）"
    if "家賃" in a or "地代" in a: return "家賃"
    if "食料品" in a or "スーパー" in a: return "食料品・スーパー"
    if "百貨店" in a: return "百貨店・買物"
    if "衣服" in a: return "衣服"
    if "家電" in a: return "家電"
    if "コンビニ" in a: return "コンビニ"
    if "ドラッグ" in a or "薬" in a: return "ドラッグ・薬"
    if "クリーニング" in a: return "クリーニング"
    if "植物" in a or "雑貨" in a: return "雑貨・その他物販"
    if "YouTube" in a or "サブスク" in a: return "サブスク"
    if any(k in a for k in ["オリコ","アプラス","Paidy","ペイディ","ペイデイ"]): return "個人カード返済"
    if "奨学金" in a: return "奨学金返済"
    if "現金" in a or "ATM" in a: return "現金引出・現金支出"
    if "未分類" in a: return "未分類（生活費）"
    return "その他個人支出"

# ---- 子内訳（項目）ラベル ----
_ITEM_MAP = [
    ("Google Workspace","Google Workspace"),("Claude","生成AI（Claude）"),("生成AI","生成AI"),
    ("YouTube","YouTube Premium"),("Zoom","Zoom"),("freee","freee"),("Google Cloud","Google Cloud"),
    ("携帯","携帯電話"),("サーバー","サーバー/ドメイン"),("03","IP電話（03plus）"),
    ("税理士","税理士 顧問料"),("BIZ SPOT","コワーキング（BIZ SPOT）"),("IIOFFICE","コワーキング（IIOFFICE）"),
    ("レゾナンス","バーチャルオフィス（レゾナンス）"),("バーチャルオフィス","バーチャルオフィス"),
    ("電気","電気"),("水道","水道"),("ガス","ガス"),
    ("Pay-easy","社会保険料（Pay-easy）"),("社会保険","社会保険料"),("役員報酬","役員報酬（定期）"),
    ("外貨積立","外貨積立（USD・650+200→7月〜200）"),("家賃保証","家賃保証"),("ヤチン","家賃（社宅収納179,550）"),("家賃","家賃（個人住居）"),
    ("奨学金","奨学金返済"),("オリコ","オリコ（リボ+分割）"),("アプラス","アプラス"),("DC","DCカード"),
    ("スマホ","スマホ分割（Google・SplitIt 月3,012）"),("SPLITIT","スマホ分割（Google・SplitIt）"),
    ("ペイデイ","Paidy（後払い）"),("ペイディ","Paidy（後払い）"),("Paidy","Paidy（後払い）"),
    ("タクシー","タクシー/配車"),("宿泊","宿泊"),("航空","航空/旅行"),("空港","空港"),("交通系","交通系IC(PASMO/Suica)"),
    ("シェアサイクル","シェアサイクル(LUUP)"),("電車","電車/新幹線"),
    ("社宅","社宅・不動産"),
]
def item_name(acc):
    for kw, label in _ITEM_MAP:
        if kw in acc: return label
    m = re.search(r"[(（]([^)）・]+)", acc)  # 括弧内の主要語（人名等）
    if m and m.group(1) not in ("人名", "個人"):
        return m.group(1)[:20]
    s = acc.split("・")[-1].split("（")[0].split("(")[0].strip()
    return s[:20] or acc[:20]

CORP_FIX = ["役員報酬","給料手当","法定福利費","通信費","ソフトウェア利用料","支払報酬料","地代家賃","水道光熱費"]
CORP_VAR = ["外注費","旅費交通費","会議費","接待交際費","広告宣伝費","消耗品費","新聞図書費","租税公課","支払手数料","雑費（現金支出）","その他経費"]
KOJIN_FIX = ["外貨積立（USD貯蓄）","家賃","奨学金返済","個人カード返済","スマホ分割（Google）","サブスク"]
KOJIN_VAR = ["食料品・スーパー","コンビニ","百貨店・買物","衣服","家電","ドラッグ・薬","クリーニング","雑貨・その他物販","現金引出・現金支出","未分類（生活費）","その他個人支出"]

# ---- 集計（親→子→月） ----
ci = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))            # 法人収入 [y][取引先][m]
ce = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))  # 法人経費 [y][親][子][m]
ki = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))            # 個人収入 [y][項目][m]
ke = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))  # 個人支出 [y][親][子][m]
for r in rows:
    y = yr(r); m = mo(r); seg = r["segment"]
    if y not in ("2025", "2026") or m == 0: continue
    if seg == "法人" and r["flow"] == "入金":
        ci[y][kamoku_corp_inc(r["account"])][m] += A(r)
    elif seg == "法人" and r["flow"] == "出金":
        p = kamoku_corp_exp(r["account"]); ch = item_name(r["account"])
        if "携帯電話" in r["account"]:  # ドコモは月2回線。金額で分割
            ch = "携帯① ドコモ(約5千)" if A(r) < 6000 else "携帯② ドコモ(約8千)"
        ce[y][p][ch][m] += A(r)
    elif seg == "役員報酬":
        ce[y]["役員報酬"]["役員報酬（定期）"][m] += A(r)
    elif seg == "役員報酬受取":
        ki[y]["役員報酬 受取"][m] += A(r)
    elif seg == "個人候補" and r["flow"] == "入金":
        ki[y]["その他収入"][m] += A(r)
    elif seg == "個人候補" and r["flow"] == "出金":
        p = kamoku_kojin(r["account"]); ke[y][p][item_name(r["account"])][m] += A(r)

# オリコ子を契約A(リボ・月7,780)/契約B(分割+変動)に分割
for y in ("2025", "2026"):
    cards = ke[y].get("個人カード返済", {})
    ori_key = next((k for k in cards if "オリコ" in k), None)
    if ori_key:
        ori = cards.pop(ori_key)
        for m, v in ori.items():
            a = min(7780, v); b = v - a
            cards["オリコ 契約A（リボ・月7,780）"][m] += a
            if b: cards["オリコ 契約B（分割24回ほか）"][m] += b

# 2026/7〜 外貨積立は200USD=月25,000のたたき台（実績の無い月のみ）
for m in range(7, 13):
    if ke["2026"]["外貨積立（USD貯蓄）"]["外貨積立（USD・650+200→7月〜200）"].get(m, 0) == 0:
        ke["2026"]["外貨積立（USD貯蓄）"]["外貨積立（USD・650+200→7月〜200）"][m] = 25000
# BizConfort（本人申告：法人の会議費）2026/7〜 月22,000、7月のみ+5,000=27,000（想定）
for m in range(7, 13):
    ce["2026"]["会議費"]["BizConfort（2026/7〜・月22,000）"][m] += 27000 if m == 7 else 22000

# 売上の部 表示順（取引先別・金額降順→雑収入→その他）
_ik = set(ci["2025"].keys()) | set(ci["2026"].keys())
def _it(k): return sum(ci["2025"][k].values()) + sum(ci["2026"][k].values())
CORP_INC_ORDER = sorted([k for k in _ik if k.startswith("売上")], key=lambda k: -_it(k)) + [k for k in ["雑収入","その他収入"] if k in _ik]

# ==== xlsx ====
wb = openpyxl.Workbook()
thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
H = PatternFill("solid", fgColor="305496"); HF = Font(color="FFFFFF", bold=True)
SEC = PatternFill("solid", fgColor="D6E4F0"); SECF = Font(bold=True)
TOT = PatternFill("solid", fgColor="FCE4D6"); TOTF = Font(bold=True)
GREEN = PatternFill("solid", fgColor="C6E0B4"); GREY = PatternFill("solid", fgColor="E7E6E6")
PFILL = PatternFill("solid", fgColor="DDEBF7")  # 親科目（内訳の合計）行

def header(ws, title):
    ws.cell(1, 1, title).font = Font(bold=True, size=13, color="305496")
    ws.cell(2, 1, "金額：円／利用日ベース／親科目＝太字・子内訳＝インデント／Amazon・少額要確認は未反映").font = Font(size=9, color="808080")
    hdr = ["勘定科目／内訳"] + [f"{m}月" for m in MONTHS] + ["年計"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(3, j, h); c.fill = H; c.font = HF; c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 30
    for j in range(2, 15): ws.column_dimensions[chr(64 + j)].width = 10
    ws.freeze_panes = "B4"

def leaf_block(ws, r0, title, order, data_y, total_label):
    """子のない単層ブロック（売上・個人収入）。"""
    ws.cell(r0, 1, title).font = SECF
    for c in range(1, 15): ws.cell(r0, c).fill = SEC
    r = r0 + 1; col = defaultdict(int)
    for k in order:
        md = data_y.get(k, {})
        if not md or sum(md.values()) == 0: continue
        ws.cell(r, 1, "　" + k); yt = 0
        for i, m in enumerate(MONTHS):
            v = md.get(m, 0); ws.cell(r, 2 + i, v).number_format = "#,##0"; yt += v; col[m] += v
        ws.cell(r, 14, yt).number_format = "#,##0"; ws.cell(r, 14).font = Font(bold=True)
        r += 1
    gt = total_line(ws, r, total_label, col, TOT); return r, col, gt

def hier_block(ws, r0, title, parent_order, data_y, total_label):
    """親→子の階層ブロック（販管費・個人支出）。親=太字合計、子=インデント。"""
    ws.cell(r0, 1, title).font = SECF
    for c in range(1, 15): ws.cell(r0, c).fill = SEC
    r = r0 + 1; col = defaultdict(int)
    for p in parent_order:
        children = data_y.get(p, {})
        ptot = defaultdict(int)
        for ch, md in children.items():
            for m, v in md.items(): ptot[m] += v
        if sum(ptot.values()) == 0: continue
        # 親行（＝子内訳の合計・太字＋色）
        ws.cell(r, 1, p).font = Font(bold=True)
        yt = 0
        for i, m in enumerate(MONTHS):
            ws.cell(r, 2 + i, ptot[m]).number_format = "#,##0"; yt += ptot[m]; col[m] += ptot[m]
        ws.cell(r, 14, yt).number_format = "#,##0"; ws.cell(r, 14).font = Font(bold=True)
        for c in range(1, 15): ws.cell(r, c).fill = PFILL
        r += 1
        # 子行（内訳・複数あるときのみ）
        if len(children) > 1:
            for ch in sorted(children, key=lambda c: -sum(children[c].values())):
                if sum(children[ch].values()) == 0: continue
                ws.cell(r, 1, "　　└ " + ch).font = Font(size=9, color="595959")
                for i, m in enumerate(MONTHS):
                    v = children[ch].get(m, 0)
                    cc = ws.cell(r, 2 + i, v); cc.number_format = "#,##0"; cc.font = Font(size=9, color="595959")
                ct = ws.cell(r, 14, sum(children[ch].values())); ct.number_format = "#,##0"; ct.font = Font(size=9, color="595959")
                r += 1
    gt = total_line(ws, r, total_label, col, TOT); return r, col, gt

def total_line(ws, r, label, col, fill, green=False):
    ws.cell(r, 1, label); gt = 0
    for i, m in enumerate(MONTHS):
        ws.cell(r, 2 + i, col.get(m, 0)).number_format = "#,##0"; gt += col.get(m, 0)
    ws.cell(r, 14, gt).number_format = "#,##0"
    f = GREEN if green else fill
    for c in range(1, 15): ws.cell(r, c).fill = f; ws.cell(r, c).font = Font(bold=True)
    return gt

def _add(a, b): return {m: a.get(m, 0) + b.get(m, 0) for m in MONTHS}

def corp_sheet(ws, y):
    header(ws, f"ITERRA株式会社 法人収支 月次（{y}年・実績）")
    r = 4
    r, inc, _ = leaf_block(ws, r, "【売上の部】", CORP_INC_ORDER, ci.get(y, {}), "売上・収入 合計"); r += 2
    r, fix, _ = hier_block(ws, r, "【販管費 ─ 固定費】", CORP_FIX, ce.get(y, {}), "固定費 小計"); r += 2
    r, var, _ = hier_block(ws, r, "【販管費 ─ 変動費】", CORP_VAR, ce.get(y, {}), "変動費 小計"); r += 1
    exp = _add(fix, var)
    total_line(ws, r, "販管費 合計", exp, GREY); r += 1
    total_line(ws, r, "営業利益", {m: inc.get(m, 0) - exp.get(m, 0) for m in MONTHS}, None, green=True)

def kojin_sheet(ws, y):
    header(ws, f"個人（石田優輝）収支 月次（{y}年・実績{'＋7〜12月想定' if y == '2026' else ''}）")
    r = 4
    r, inc, _ = leaf_block(ws, r, "【収入】", ["役員報酬 受取", "その他収入"], ki.get(y, {}), "収入 合計"); r += 2
    r, fix, _ = hier_block(ws, r, "【支出 ─ 固定費】", KOJIN_FIX, ke.get(y, {}), "固定費 小計"); r += 2
    r, var, _ = hier_block(ws, r, "【支出 ─ 変動費】", KOJIN_VAR, ke.get(y, {}), "変動費 小計"); r += 1
    exp = _add(fix, var)
    total_line(ws, r, "支出 合計", exp, GREY); r += 1
    total_line(ws, r, "個人収支", {m: inc.get(m, 0) - exp.get(m, 0) for m in MONTHS}, None, green=True)

ws1 = wb.active; ws1.title = "法人_月次2025"; corp_sheet(ws1, "2025")
corp_sheet(wb.create_sheet("法人_月次2026"), "2026")
kojin_sheet(wb.create_sheet("個人_月次2025"), "2025")
kojin_sheet(wb.create_sheet("個人_月次2026"), "2026")

# ---- 統合サマリ（法人＋個人を1表に・列＝年×月）----
wsS = wb.create_sheet("統合サマリ")
M25 = list(range(1, 13)); M26 = list(range(1, 7))
NC = 1 + len(M25) + 1 + len(M26) + 1  # 科目 + 2025月 + 年計 + 2026月 + 計
wsS.column_dimensions["A"].width = 32
from openpyxl.utils import get_column_letter
for c in range(2, NC + 1): wsS.column_dimensions[get_column_letter(c)].width = 8.5
wsS.cell(1, 1, "法人＋個人 統合収支（1表・列＝年×月）").font = Font(bold=True, size=13, color="305496")
wsS.cell(2, 1, "親科目＝太字／子内訳＝インデント。利用日ベース。2026は実績1〜6月（外貨積立7月以降の想定は含まない）。").font = Font(size=9, color="808080")
# ヘッダ（年・月の2段）
c25s, c25t, c26s, c26t = 2, 2 + len(M25), 2 + len(M25) + 1, NC
wsS.merge_cells(start_row=3, start_column=c25s, end_row=3, end_column=c25t)
wsS.cell(3, c25s, "2025年").alignment = Alignment(horizontal="center"); wsS.cell(3, c25s).fill = H; wsS.cell(3, c25s).font = HF
wsS.merge_cells(start_row=3, start_column=c26s, end_row=3, end_column=c26t)
wsS.cell(3, c26s, "2026年（1〜6月）").alignment = Alignment(horizontal="center"); wsS.cell(3, c26s).fill = H; wsS.cell(3, c26s).font = HF
wsS.cell(4, 1, "勘定科目／内訳").fill = H; wsS.cell(4, 1).font = HF
colmap = {}  # (year,month or 'tot') -> col index
cc = 2
for m in M25: wsS.cell(4, cc, f"{m}月"); colmap[("2025", m)] = cc; cc += 1
wsS.cell(4, cc, "年計"); colmap[("2025", "tot")] = cc; cc += 1
for m in M26: wsS.cell(4, cc, f"{m}月"); colmap[("2026", m)] = cc; cc += 1
wsS.cell(4, cc, "計"); colmap[("2026", "tot")] = cc
for c in range(2, NC + 1):
    wsS.cell(4, c).fill = H; wsS.cell(4, c).font = Font(color="FFFFFF", bold=True, size=9); wsS.cell(4, c).alignment = Alignment(horizontal="center")
wsS.freeze_panes = "B5"

rS = [5]
def srend(label, md25, md26, style=""):
    """style: ''=child, 'parent','sec','tot','grey','green'."""
    r = rS[0]
    lab = ("　　└ " + label) if style == "" else label
    c1 = wsS.cell(r, 1, lab)
    small = (style == "")
    fnt = Font(size=9, color="595959") if small else (Font(bold=True) if style in ("parent", "tot", "grey", "green") else (SECF if style == "sec" else None))
    if fnt: c1.font = fnt
    if style != "sec":
        for (y, ms) in (("2025", M25), ("2026", M26)):
            md = md25 if y == "2025" else md26
            tot = 0
            for m in ms:
                v = md.get(m, 0) if md else 0
                cell = wsS.cell(r, colmap[(y, m)], v); cell.number_format = "#,##0"
                if small: cell.font = Font(size=9, color="595959")
                elif style in ("parent", "tot", "grey", "green"): cell.font = Font(bold=True)
                tot += v
            tcell = wsS.cell(r, colmap[(y, "tot")], tot); tcell.number_format = "#,##0"
            tcell.font = Font(bold=True, size=9) if small else Font(bold=True)
    fill = {"sec": SEC, "tot": TOT, "grey": GREY, "green": GREEN, "parent": PFILL}.get(style)
    if fill:
        for c in range(1, NC + 1): wsS.cell(r, c).fill = fill
    rS[0] += 1

def msum(*mds):
    out = defaultdict(int)
    for md in mds:
        for m, v in (md or {}).items(): out[m] += v
    return out
def mdiff(a, b):
    out = defaultdict(int)
    for m in set(a) | set(b): out[m] = a.get(m, 0) - b.get(m, 0)
    return out

def hier_rows(parent_order):
    """法人/個人 経費の親→子行を出力し、(合計md25, 合計md26)を返す。"""
    t25, t26 = defaultdict(int), defaultdict(int)
    for p in parent_order:
        ch25 = ce["2025"].get(p) if parent_order is CORP_FIX or parent_order is CORP_VAR else ke["2025"].get(p)
        ch26 = ce["2026"].get(p) if parent_order is CORP_FIX or parent_order is CORP_VAR else ke["2026"].get(p)
        ch25 = ch25 or {}; ch26 = ch26 or {}
        p25 = msum(*ch25.values()); p26 = msum(*ch26.values())
        if sum(p25.values()) == 0 and sum(p26.values()) == 0: continue
        srend(p, p25, p26, "parent")
        for m, v in p25.items(): t25[m] += v
        for m, v in p26.items(): t26[m] += v
        keys = set(ch25) | set(ch26)
        if len(keys) > 1:
            for k in sorted(keys, key=lambda k: -sum((ch25.get(k) or {}).values())):
                srend(k, ch25.get(k, {}), ch26.get(k, {}), "")
    return t25, t26

# 【法人】
srend("【法人】売上の部", None, None, "sec")
si25, si26 = defaultdict(int), defaultdict(int)
for k in CORP_INC_ORDER:
    md25 = ci["2025"].get(k, {}); md26 = ci["2026"].get(k, {})
    if sum(md25.values()) + sum(md26.values()) == 0: continue
    srend(k, md25, md26, "")
    si25 = msum(si25, md25); si26 = msum(si26, md26)
srend("売上・収入 合計", si25, si26, "tot")
srend("【法人】販管費 ─ 固定費", None, None, "sec")
cf25, cf26 = hier_rows(CORP_FIX); srend("固定費 小計", cf25, cf26, "tot")
srend("【法人】販管費 ─ 変動費", None, None, "sec")
cv25, cv26 = hier_rows(CORP_VAR); srend("変動費 小計", cv25, cv26, "tot")
exp25 = msum(cf25, cv25); exp26 = msum(cf26, cv26)
srend("販管費 合計", exp25, exp26, "grey")
op25 = mdiff(si25, exp25); op26 = mdiff(si26, exp26)
srend("営業利益（法人）", op25, op26, "green")
# 【個人】
srend("【個人】収入", None, None, "sec")
pi25, pi26 = defaultdict(int), defaultdict(int)
for k in ("役員報酬 受取", "その他収入"):
    md25 = ki["2025"].get(k, {}); md26 = ki["2026"].get(k, {})
    if sum(md25.values()) + sum(md26.values()) == 0: continue
    srend(k, md25, md26, "")
    pi25 = msum(pi25, md25); pi26 = msum(pi26, md26)
srend("収入 合計", pi25, pi26, "tot")
srend("【個人】支出 ─ 固定費", None, None, "sec")
kf25, kf26 = hier_rows(KOJIN_FIX); srend("固定費 小計", kf25, kf26, "tot")
srend("【個人】支出 ─ 変動費", None, None, "sec")
kv25, kv26 = hier_rows(KOJIN_VAR); srend("変動費 小計", kv25, kv26, "tot")
pe25 = msum(kf25, kv25); pe26 = msum(kf26, kv26)
srend("支出 合計", pe25, pe26, "grey")
pb25 = mdiff(pi25, pe25); pb26 = mdiff(pi26, pe26)
srend("個人収支", pb25, pb26, "green")
# 【統合】純収支（法人営業利益＋個人収支＝役員報酬は相殺）
srend("【統合】純収支（法人営業利益＋個人収支）", msum(op25, pb25), msum(op26, pb26), "green")
# 収支外（年計のみ・月別なし）
srend("【収支外・役員勘定（損益外）】", None, None, "sec")
def sy(seg, y): return sum(A(r) for r in rows if r["segment"] == seg and yr(r) == y)
for lab, seg in (("役員貸付（法人→個人）", "役員貸付"), ("役員借入（個人→法人）", "役員借入"),
                 ("除外計（カード決済振替・借入返済・手許現金・自社間振替・各種戻し）", "除外"), ("要確認 残（Amazon保留含む）", "要確認")):
    r = rS[0]
    wsS.cell(r, 1, lab)
    wsS.cell(r, colmap[("2025", "tot")], sy(seg, "2025")).number_format = "#,##0"
    wsS.cell(r, colmap[("2026", "tot")], sy(seg, "2026")).number_format = "#,##0"
    rS[0] += 1

_unused_old = """
wsS_old = wb.create_sheet("統合サマリ_old")
wsS.column_dimensions["A"].width = 40
for j in (2, 3): wsS.column_dimensions[chr(64 + j)].width = 15
wsS.cell(1, 1, "法人＋個人 統合サマリ（年計・内訳付き・円・実績）").font = Font(bold=True, size=13, color="305496")
wsS.cell(2, 1, "親科目＝太字／子内訳＝インデント。2026は1〜6月実績（外貨積立7月以降の想定は含まない）。").font = Font(size=9, color="808080")
for j, h in enumerate(["勘定科目／内訳", "2025", "2026(1〜6月)"], 1):
    c = wsS.cell(3, j, h); c.fill = H; c.font = HF; c.alignment = Alignment(horizontal="center")
rS = [4]
# 月範囲：2025=1〜12、2026=1〜6（実績）
def ytot(md, y): return sum(md.get(m, 0) for m in (range(1, 13) if y == "2025" else range(1, 7)))
def ptot(children, y): return sum(ytot(md, y) for md in children.values())
def srow(label, v25, v26, fill=None, bold=False, child=False, money=True):
    c1 = wsS.cell(rS[0], 1, ("　　└ " if child else "") + label)
    if child: c1.font = Font(size=9, color="595959")
    elif bold: c1.font = Font(bold=True)
    for col, v in ((2, v25), (3, v26)):
        cc = wsS.cell(rS[0], col, v)
        if money: cc.number_format = "#,##0"
        if child: cc.font = Font(size=9, color="595959")
        elif bold: cc.font = Font(bold=True)
    if fill:
        for c in range(1, 4): wsS.cell(rS[0], c).fill = fill
    rS[0] += 1
def sec(label):
    wsS.cell(rS[0], 1, label).font = SECF
    for c in range(1, 4): wsS.cell(rS[0], c).fill = SEC
    rS[0] += 1
def hier(parent_order, dataD):
    f = [0, 0]
    for p in parent_order:
        ch = dataD["2025"].get(p, {}); ch26 = dataD["2026"].get(p, {})
        p25 = ptot(ch, "2025"); p26 = ptot(ch26, "2026")
        if p25 == 0 and p26 == 0: continue
        srow(p, p25, p26, bold=True); f[0] += p25; f[1] += p26
        keys = set(ch) | set(ch26)
        if len(keys) > 1:
            for k in sorted(keys, key=lambda k: -(ptot({0: ch.get(k, {})}, "2025"))):
                srow(k, ytot(ch.get(k, {}), "2025"), ytot(ch26.get(k, {}), "2026"), child=True)
    return f

# 【法人】
sec("【法人】売上の部")
si = [0, 0]
for k in CORP_INC_ORDER:
    v25 = ytot(ci["2025"].get(k, {}), "2025"); v26 = ytot(ci["2026"].get(k, {}), "2026")
    if v25 or v26: srow(k, v25, v26, child=True); si[0] += v25; si[1] += v26
srow("売上・収入 合計", si[0], si[1], TOT, True)
sec("【法人】販管費 ─ 固定費")
cf = hier(CORP_FIX, ce)
srow("固定費 小計", cf[0], cf[1], TOT, True)
sec("【法人】販管費 ─ 変動費")
cv = hier(CORP_VAR, ce)
srow("変動費 小計", cv[0], cv[1], TOT, True)
srow("販管費 合計", cf[0] + cv[0], cf[1] + cv[1], GREY, True)
srow("営業利益", si[0] - cf[0] - cv[0], si[1] - cf[1] - cv[1], GREEN, True)
rS[0] += 1
# 【個人】
sec("【個人】収入")
pi = [0, 0]
for k in ("役員報酬 受取", "その他収入"):
    v25 = ytot(ki["2025"].get(k, {}), "2025"); v26 = ytot(ki["2026"].get(k, {}), "2026")
    if v25 or v26: srow(k, v25, v26, child=True); pi[0] += v25; pi[1] += v26
srow("収入 合計", pi[0], pi[1], TOT, True)
sec("【個人】支出 ─ 固定費")
kf = hier(KOJIN_FIX, ke)
srow("固定費 小計", kf[0], kf[1], TOT, True)
sec("【個人】支出 ─ 変動費")
kv = hier(KOJIN_VAR, ke)
srow("変動費 小計", kv[0], kv[1], TOT, True)
srow("支出 合計", kf[0] + kv[0], kf[1] + kv[1], GREY, True)
srow("個人収支", pi[0] - kf[0] - kv[0], pi[1] - kf[1] - kv[1], GREEN, True)
rS[0] += 1
# 収支外・役員勘定
def sy(seg, y):
    return sum(A(r) for r in rows if r["segment"] == seg and yr(r) == y)
sec("【収支外・役員勘定（損益外）】")
srow("役員貸付（法人→個人）", sy("役員貸付", "2025"), sy("役員貸付", "2026"))
srow("役員借入（個人→法人）", sy("役員借入", "2025"), sy("役員借入", "2026"))
srow("除外計（カード決済振替・借入返済・手許現金・自社間振替・各種戻し）", sy("除外", "2025"), sy("除外", "2026"))
srow("要確認 残（Amazon保留含む）", sy("要確認", "2025"), sy("要確認", "2026"))
"""

out = os.path.join(BASE, "収支実績_法人個人_2025-2026.xlsx")
wb.save(out)
print("生成:", out, "／ シート:", wb.sheetnames)
